import streamlit as st
import pandas as pd
import pdfplumber
import re
from datetime import datetime
import io
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

def extract_ledger_data(pdf_file):
    """Extracts transaction data from an uploaded PDF file object."""
    transactions = []
    party_name = "Unknown Party"
    
    with pdfplumber.open(pdf_file) as pdf:
        text = ""
        for page in pdf.pages:
            extracted = page.extract_text()
            if extracted:
                text += extracted + "\n"
            
    lines = text.split('\n')
    
    for i, line in enumerate(lines):
        if "Ledger Account" in line and i > 0:
            party_name = lines[i-1].strip()
            break

    date_pattern = re.compile(r'(\d{1,2}-[A-Za-z]{3}-\d{2})') 
    current_date = None
    
    for line in lines:
        date_match = date_pattern.search(line)
        if date_match:
            current_date = date_match.group(1)
            date_obj = datetime.strptime(current_date, '%d-%b-%y')
            month_year = date_obj.strftime('%B %Y')
        
        if not current_date:
            continue
            
        clean_line = line.replace('|', '').replace(',', '').strip()
        
        if "Opening Balance" in clean_line:
            numbers = re.findall(r'\d+\.\d{2}', clean_line)
            if numbers:
                transactions.append({"Party Name": party_name, "Month": month_year, "Type": "Opening Balance", "Amount": float(numbers[-1])})
                
        elif "Sales" in clean_line and "AH/" in clean_line:
            numbers = re.findall(r'\d+\.\d{2}', clean_line)
            if numbers:
                transactions.append({"Party Name": party_name, "Month": month_year, "Type": "Sale Debit", "Amount": float(numbers[-1])})
                
        elif "Credit Note" in clean_line and "CN" in clean_line:
            numbers = re.findall(r'\d+\.\d{2}', clean_line)
            if numbers:
                transactions.append({"Party Name": party_name, "Month": month_year, "Type": "Sale Credit", "Amount": float(numbers[-1])})
                
        elif "Receipt" in clean_line:
            numbers = re.findall(r'\d+\.\d{2}', clean_line)
            if numbers:
                transactions.append({"Party Name": party_name, "Month": month_year, "Type": "Receipt", "Amount": float(numbers[-1])})

    return transactions

def create_excel(all_data):
    """Processes the raw data and generates a styled Excel file in memory."""
    df = pd.DataFrame(all_data)
    final_rows = []
    
    months_order = ['April 2026', 'May 2026', 'June 2026', 'July 2026', 'August 2026'] 
    
    for party, party_data in df.groupby("Party Name"):
        running_outstanding = 0.0
        party_data['Month_Cat'] = pd.Categorical(party_data['Month'], categories=months_order, ordered=True)
        party_data = party_data.sort_values('Month_Cat')
        
        for month, month_data in party_data.groupby("Month_Cat", observed=True):
            if month_data.empty:
                continue
                
            opening_bal = month_data[month_data['Type'] == 'Opening Balance']['Amount'].sum()
            sale_debit = month_data[month_data['Type'] == 'Sale Debit']['Amount'].sum()
            sale_credit = month_data[month_data['Type'] == 'Sale Credit']['Amount'].sum()
            receipt = month_data[month_data['Type'] == 'Receipt']['Amount'].sum()
            
            if opening_bal > 0:
                running_outstanding += opening_bal
                
            running_outstanding = (running_outstanding + sale_debit) - (sale_credit + receipt)
            
            final_rows.append({
                "Party Name": party,
                "Month": month,
                "Total sale offline debit": sale_debit,
                "Sale offline credit": sale_credit,
                "Payment receipt": receipt,
                "Total Outstanding": running_outstanding
            })
            
    final_df = pd.DataFrame(final_rows)
    final_df['Month_Cat'] = pd.Categorical(final_df['Month'], categories=months_order, ordered=True)
    final_df = final_df.sort_values(by=['Month_Cat', 'Party Name']).drop(columns=['Month_Cat'])
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        final_df.to_excel(writer, index=False, sheet_name='Consolidated Ledger')
        
        workbook = writer.book
        worksheet = writer.sheets['Consolidated Ledger']
        
        # --- APPLYING EXCEL STYLES ---
        # Define Fonts and Borders
        base_font = Font(name='Times New Roman', size=12)
        bold_font = Font(name='Times New Roman', size=12, bold=True)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Apply styles to headers (Row 1)
        for col_num, value in enumerate(final_df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.font = bold_font
            cell.border = thin_border
            
        # Apply styles to data rows (Row 2 onwards)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.font = base_font
                cell.border = thin_border
                # Number formatting for currency columns
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                
        # Auto-adjust column widths based on maximum length of text in each column
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Add padding to make the columns comfortably wide
            adjusted_width = (max_length + 6) 
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    return buffer.getvalue()

# --- STATE MANAGEMENT ---
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0
if "processing_complete" not in st.session_state:
    st.session_state.processing_complete = False
if "excel_data" not in st.session_state:
    st.session_state.excel_data = None

def reset_app():
    """Callback function triggered when the download button is clicked."""
    st.session_state.uploader_key += 1 
    st.session_state.processing_complete = False
    st.session_state.excel_data = None

# --- WEB APP INTERFACE ---
st.set_page_config(page_title="Ledger Processor", layout="centered")
st.title("📊 Ledger PDF to Excel Converter")
st.markdown("Please Upload PDFs")

uploaded_files = st.file_uploader(
    "Upload PDF Files", 
    type=["pdf"], 
    accept_multiple_files=True, 
    key=f"uploader_{st.session_state.uploader_key}"
)

# Flow Logic
if uploaded_files:
    if not st.session_state.processing_complete:
        st.info(f"You have uploaded {len(uploaded_files)} PDF(s).")
        st.write("Do you want to process them?")
        
        if st.button("Yes"):
            all_extracted_data = []
            progress_bar = st.progress(0)
            
            for i, file in enumerate(uploaded_files):
                try:
                    pdf_data = extract_ledger_data(file)
                    all_extracted_data.extend(pdf_data)
                except Exception as e:
                    st.error(f"Error processing {file.name}: {e}")
                
                progress_bar.progress((i + 1) / len(uploaded_files))
                
            if all_extracted_data:
                st.session_state.excel_data = create_excel(all_extracted_data)
                st.session_state.processing_complete = True
                st.rerun()
            else:
                st.warning("No transaction data could be extracted from the uploaded PDFs.")

    if st.session_state.processing_complete and st.session_state.excel_data:
        st.success("✅ Files processed successfully!")
        st.download_button(
            label="⬇️ Download Consolidated Excel",
            data=st.session_state.excel_data,
            file_name="Consolidated_Ledger.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            on_click=reset_app
        )
